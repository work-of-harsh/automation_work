import openpyxl                   #library to read and write in excel
import json                      #library to read and write in json

class testData:                               #all data belonging to a particular page will go to a specific class in "test_data" file.
    #providing data directly from variable.

    #we have two dictionaries in list, so tc will run total 8 times.4-4 for each dictionary set.
    page_data=[{'mobileName':'Samsung Note 8','entered_country_name':'ind','country_presence':'India'},{'mobileName':'Samsung Note 8','entered_country_name':'China','country_presence':'China'}]



    # providing data from excel.
    @staticmethod          #declaring static method to call the method directly via class without object creation.
    def exceldata(testcase_name):  #static methods don't use self.
        book = openpyxl.load_workbook("C:\\Users\\Lenovo\\Desktop\\Python\\GIT Projects\\Python\\Python_Selenium_with_PyTest_framework\\Test_Data\\Excel_data.xlsx")  # importing excel file
        sheet = book.active  # selecting active sheet
        list= []
        d={}
        for i in range(1,sheet.max_row+1):        #in excel row number starts from 1.
            if sheet.cell(row=i,column =1).value == testcase_name:      #will enter the loop only for row having "Test_Case2".
                for j in range(2,sheet.max_column+1):
                    d[sheet.cell(row=1,column =j).value] =sheet.cell(row=i,column =j).value   #extracting column values for "Test_Case2" row into dictionary format.
                list.append(d)         #fixture accepts dictionary in list.
        return list



    # providing data from json.
    @staticmethod
    def jsondata(testcase_name):
        with open("C:\\Users\Lenovo\Desktop\Python\GIT Projects\Python\Python_Selenium_with_PyTest_framework\Test_Data\Test_Data.json") as file:
            data =json.load(file)
        test_case_list = data[testcase_name]     #each testcase in json file has dictionary wrapped in list.
        print(test_case_list)
        return test_case_list

